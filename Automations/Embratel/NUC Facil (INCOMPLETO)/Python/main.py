import os
import pyodbc
import openpyxl
import warnings
import traceback
import subprocess
import pandas as pd
import tkinter as tk

from tkinter import ttk
from tkinter import messagebox

from time import time
from datetime import datetime as dt
from os.path import join, abspath, dirname
from openpyxl.utils.dataframe import dataframe_to_rows

# Suprimir todos os avisos (não recomendado a menos que seja absolutamente necessário)
# warnings.filterwarnings("ignore")

# Suprimir um tipo específico de aviso
warnings.filterwarnings("ignore", category=DeprecationWarning)

# Suprimir um aviso específico por mensagem
warnings.filterwarnings(
    "ignore",
    message="alguma mensagem de aviso específica")

# Se você deseja restaurar os avisos posteriormente, você pode usar:
# warnings.resetwarnings()

pd.set_option('mode.chained_assignment', None)


class DataframeIsEmpty(Exception):
    pass


class FileNotRecognized(Exception):
    pass


class ExcelIsOpened(Exception):
    pass


class Main_Screen:
    def __init__(self, root):
        self.now = dt.now()
        self.today = self.now.date()
        self.day = self.now.day
        self.month = self.now.month
        self.year = self.now.year

        self.current_path = os.path.dirname(__file__)
        os.chdir(self.current_path)

        self.father_path = abspath(join(self.current_path, '..'))
        self.convert_path = join(self.current_path, 'CONVERTER')
        self.destin_path = join(self.current_path, 'CONVERTIDO')
        self.nuc_path = join(self.father_path, 'NUC_Facil')

        self.root = root
        self.root.title("Atualizar SNU - Embratel")
        self.root.geometry('400x200')
        self.root.resizable(width=False, height=False)
        self.root.iconbitmap(join(self.current_path, 'file.ico'))

        self.main_pack = ttk.Frame(root)
        self.main_pack.pack()

        self.convert_pack = ttk.Frame(self.main_pack)
        self.convert_pack.grid(row=0, column=0, padx=10, pady=5)

        self.year_label = ttk.Label(
            self.convert_pack, text="Ano de referência:", )
        self.year_label.pack(padx=10, pady=5)
        self.year_entry = ttk
        self.year_entry.pack(padx=10, pady=5)
        self.month_label = ttk.Label(self.convert_pack, text="Data de Fim:")
        self.month_label.pack(padx=10, pady=5)
        self.month_entry = DateEntry(
            self.convert_pack, date_pattern="dd/mm/yyyy")
        self.month_entry.pack(padx=10, pady=5)
        self.update_button = ttk.Button(
            self.convert_pack,
            text="Criar planilha",
            command=self.merge_files)
        self.update_button.pack(padx=10, pady=5)

        self.options_pack = ttk.Frame(self.main_pack)
        self.options_pack.grid(row=0, column=1, padx=10, pady=5)

        self.open_convert_path_button = ttk.Button(
            self.options_pack,
            text='Abrir pasta de arquivos a converter',
            command=self.open_convert_files_path)
        self.open_convert_path_button.pack(padx=10, pady=5)
        self.open_destin_path_button = ttk.Button(
            self.options_pack,
            text='Abrir pasta de destino dos arquivos',
            command=self.open_destin_files_path)
        self.open_destin_path_button.pack(padx=10, pady=5)
        self.open_nuc_software_button = ttk.Button(
            self.options_pack,
            text='Abrir o software do NUC Fácil',
            command=self.open_nuc_software)
        self.open_nuc_software_button.pack(padx=10, pady=5)

    def show_custom_message(
            self, message: str, title: str = "Aviso", timeout: int = 5):
        custom_message_window = tk.Toplevel(root)
        custom_message_window.title(title)
        label = ttk.Label(custom_message_window, text=message)
        label.pack(padx=15, pady=15)
        custom_message_window.after(
            timeout * 1000,
            custom_message_window.destroy)  # Fecha após 10 segundos

    def merge_files(self):
        try:
            self.move_snu_file()
            init = time()
            df_datalens = self.get_data_datalens()
            self.move_snu_file()
            df = self.treat_csv_snu_files(df_datalens=df_datalens)
            self.create_worksheet(df=df)
            print(f"O processo demorou {time() - init} segundos")
        except Exception as e:
            print(traceback.print_exc())
            error_msg = f"Erro no código, por gentileza chamar o desenvolvedor responsável.\n\nErro principal: {e}\n\nErro completo:{traceback.format_exc()}"
            self.show_custom_message(
                title='Erro', message=error_msg, timeout=25)

    def open_convert_files_path(self):
        f'explorer "{self.convert_path}"'

    def open_destin_files_path(self):
        subprocess.Popen(f'explorer "{self.destin_path}"')

    def open_nuc_software(self):
        subprocess.Popen(f'"{join(self.nuc_path, "NUC_Facil.exe")}"')

    def command_sql(self):
        self.start_date = self.start_date_entry.get()
        self.end_date = self.end_date_entry.get()

        # Alternativa 1 (GAMB)
        return """
select
    cast(datahora as date) as data,
    cast(datahora as time(0)) as hora,
    origem,
    destino,
    categoria,
    durtar,
    valor_com_impostos
from inv_53_fixo
where
tipo_servico = 'VOZ - 0800'
and cast(datahora as date) like '2023-09%'"""

    def get_data_datalens(self):
        host = 'datalens-billing.database.windows.net'
        user = 'datalens'
        password = '0gxXTs0E'
        database = 'billing'

        # Conectar ao banco de dados
        with pyodbc.connect('DRIVER={SQL Server};' + f'SERVER={host};DATABASE={database};UID={user};PWD={password}') as connection:
            df_datalens = pd.read_sql(_ := self.command_sql(), connection)

        if df_datalens.empty:
            raise DataframeIsEmpty(
                f'O banco de dados do período de {self.start_date} a {self.end_date} está vazio')

        else:
            return df_datalens

    def move_snu_file(self):
        self.snu_files = [
            file for file in os.listdir(
                self.nuc_path) if "snu" in file.lower()]

        for file in self.snu_files:
            if '.CSV' in file:
                os.replace(
                    join(
                        self.nuc_path, file), join(
                        self.convert_path, file))
                print("Movido o arquivo", file)
            else:
                os.remove(join(self.nuc_path, file))
                print("Deletado o arquivo", file)

    def treat_csv_snu_files(self, df_datalens: pd.DataFrame):
        df = pd.DataFrame()

        for file in os.listdir(self.convert_path):
            print(join(self.convert_path, file))
            if '.csv' in file.lower() or '.txt' in file.lower():
                temp_df = pd.read_csv(
                    join(
                        self.convert_path,
                        file),
                    sep=";",
                    engine='pyarrow')
            elif '.xls' in file.lower():
                temp_df = pd.read_excel(join(self.convert_path, file))
            else:
                raise FileNotRecognized("Arquivo não reconhecido")

            temp_df.columns = temp_df.columns.str.strip()

            # Filtra o tipo do chamado
            temp_df = temp_df.loc[temp_df['TIPO'] == "Q"]

            temp_df['DATA_CHAMADA'] = pd.to_datetime(
                temp_df['DATA_CHAMADA'], format='%d/%m/%Y').dt.date.astype(str)
            temp_df['DURACAO'] = temp_df['DURACAO'].astype(
                str).str.replace(',', '.').astype(float)
            temp_df['VALOR'] = temp_df['VALOR'].astype(
                str).str.replace(',', '.').astype(float)

            temp_df = temp_df[['DATA_CHAMADA',
                               'HORA',
                               'TEL_CHAMADO',
                               'TEL_ORIGEM',
                               'TIPO',
                               'DURACAO',
                               'VALOR']]  # Organiza a ordem das colunas

            temp_df = temp_df.rename(columns={
                'DATA_CHAMADA': 'data',
                'HORA': 'hora',
                # o TEL_CHAMADO e o TEL_ORIGEM são invertidos devido a um erro
                # na planilha
                'TEL_CHAMADO': 'origem',
                'TEL_ORIGEM': 'destino',
                'TIPO': 'categoria',
                'DURACAO': 'durtar',
                'VALOR': 'valor_com_impostos'
            })

            df = pd.concat([df, temp_df])

        print(df.head(15).to_string())

        return pd.concat([df_datalens, df])

    def format_worksheet_styles(self, ws, final_df):
        # Defina a área de dados que você deseja formatar como uma tabela
        # Suponhamos que você deseja formatar a área de A1 até C10 como uma
        # tabela
        end_row = len(final_df.index) + 1
        end_col = 7

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 7
        ws.column_dimensions['G'].width = 22

        # Aplique o estilo de tabela à área de dados
        tab = openpyxl.worksheet.table.Table(
            displayName="Tabela_0800",
            ref=f"{ws.cell(1, 1).coordinate}:{ws.cell(end_row, end_col).coordinate}")
        style = openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleLight14", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        return ws

    def create_worksheet(self, df):  # sourcery skip: raise-from-previous-error
        wb = openpyxl.Workbook()  # Cria um novo arquivo excel
        ws = wb.active

        for r in dataframe_to_rows(
                df, index=False, header=True):  # Insere DF no excel
            ws.append(r)

        ws = self.format_worksheet_styles(ws, df)

        # Salve o arquivo Excel
        path_file = join(
            self.destin_path,
            f'Relatório Electrolux - Ligações {dt.strptime(self.end_date, "%d/%m/%Y").strftime("%Y-%m")}.xlsx')
        try:
            wb.save(path_file)
        except PermissionError:
            raise ExcelIsOpened(
                f'Verifique se o arquivo  "Relatório Electrolux - Ligações {dt.strptime(self.end_date, "%d/%m/%Y").strftime("%Y-%m")}.xlsx" encontra-se aberto')

        self.show_custom_message(
            title="Sucesso",
            message=f"O relatório foi criado em {path_file}")


if __name__ == '__main__':
    root = tk.Tk()
    app = Main_Screen(root)
    root.mainloop()
